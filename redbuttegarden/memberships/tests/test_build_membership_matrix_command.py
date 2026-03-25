from io import StringIO

from django.core.management import call_command
from openpyxl import load_workbook

from memberships.services.matrix import DEFAULT_MEMBERSHIP_MATRIX_FIXTURE_PATH
from memberships.services.recommendations import (
    DEFAULT_PRICE_FALLBACK_FORMULAS,
    DEFAULT_RECOMMENDATION_FORMULAS,
)


def test_build_membership_matrix_command_writes_expected_workbook(tmp_path):
    out_path = tmp_path / "membership_matrix.xlsx"
    stdout = StringIO()

    call_command("build_membership_matrix", out=str(out_path), stdout=stdout)

    assert out_path.exists()

    workbook = load_workbook(out_path)
    assert workbook.sheetnames == ["Matrix", "Levels", "Meta"]

    meta = {
        key: value
        for key, value in workbook["Meta"].iter_rows(values_only=True)
        if key is not None
    }
    assert meta["level_source"] == str(DEFAULT_MEMBERSHIP_MATRIX_FIXTURE_PATH)
    assert meta["downsell_1_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["downsell_1"]
    )
    assert meta["downsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
    )
    assert meta["upsell_1_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
    )
    assert meta["upsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
    )
    assert (
        meta["downsell_1_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["downsell_1"]
    )
    assert (
        meta["downsell_2_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["downsell_2"]
    )
    assert (
        meta["upsell_1_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_1"]
    )
    assert (
        meta["upsell_2_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_2"]
    )
    assert "Wrote:" in stdout.getvalue()
    assert "Rows:" in stdout.getvalue()
