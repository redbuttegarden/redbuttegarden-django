from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from memberships.services.matrix import (
    build_membership_matrix_rows,
    build_membership_matrix_workbook_bytes,
)
from memberships.services.recommendations import DEFAULT_RECOMMENDATION_FORMULAS, Level


def make_level(pk, name, cardholders, guests, tickets, price):
    return Level(
        pk=pk,
        name=name,
        cardholders_included=cardholders,
        admissions_allowed=guests,
        member_sale_ticket_allowance=tickets,
        price=Decimal(price),
    )


def make_cfg():
    return SimpleNamespace(
        cardholder_label="",
        admissions_label="",
        member_tickets_label="",
        presale_qualification_error_message_template=(
            "Please qualify for {tickets} presale tickets."
        ),
    )


def test_matrix_rows_use_supplied_formulas_instead_of_defaults():
    levels = [
        make_level(1, "Highlighted", 1, 2, 2, "100.00"),
        make_level(2, "Default Downsell", 1, 2, 0, "80.00"),
        make_level(3, "Custom Downsell", 2, 2, 0, "90.00"),
        make_level(4, "Fallback Downsell", 1, 1, 2, "70.00"),
        make_level(5, "Upsell", 1, 3, 2, "120.00"),
    ]

    rows = build_membership_matrix_rows(
        levels=levels,
        cfg=make_cfg(),
        formulas={"downsell_1": ("(C+1, G, prev(T))",)},
    )

    row = next(
        item
        for item in rows
        if item["cardholders"] == 1
        and item["admissions"] == 2
        and item["tickets"] == 2
    )
    assert row["downsell_1_name"] == "Custom Downsell"
    assert row["downsell_1_formula"] == "(C+1, G, prev(T))"


def test_matrix_workbook_bytes_include_meta_for_level_source_and_formulas():
    levels = [
        make_level(1, "Highlighted", 1, 2, 2, "100.00"),
        make_level(2, "Downsell", 1, 2, 0, "80.00"),
        make_level(3, "Upsell", 1, 3, 2, "120.00"),
    ]
    formulas = {"downsell_1": ("(C+1, G, prev(T))",)}
    rows = build_membership_matrix_rows(levels=levels, cfg=make_cfg(), formulas=formulas)

    workbook_bytes = build_membership_matrix_workbook_bytes(
        rows=rows,
        levels=levels,
        level_source="fixture:memberships/fixtures/membership_levels.json",
        formulas=formulas,
    )

    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Matrix", "Levels", "Meta"]

    meta = {
        key: value
        for key, value in workbook["Meta"].iter_rows(values_only=True)
        if key is not None
    }
    assert (
        meta["level_source"] == "fixture:memberships/fixtures/membership_levels.json"
    )
    assert "MembershipSelectorForm domain" in meta["notes"]
    assert meta["downsell_1_formulas"] == "(C+1, G, prev(T))"
    assert meta["downsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
    )
    assert meta["upsell_1_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
    )
    assert meta["upsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
    )
