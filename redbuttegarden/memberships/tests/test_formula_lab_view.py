from io import BytesIO

from django.urls import reverse

import pytest
from openpyxl import load_workbook

from memberships.models import MembershipLevel
from memberships.services.matrix import DEFAULT_MEMBERSHIP_MATRIX_FIXTURE_PATH
from memberships.services.recommendations import (
    DEFAULT_PRICE_FALLBACK_FORMULAS,
    DEFAULT_RECOMMENDATION_FORMULAS,
)


def create_level(name, cardholders, guests, tickets, price):
    return MembershipLevel.objects.create(
        name=name,
        cardholders_included=cardholders,
        admissions_allowed=guests,
        member_sale_ticket_allowance=tickets,
        price=price,
        active=True,
    )


def default_fallback_post_data():
    return {
        "downsell_1_price_fallback": DEFAULT_PRICE_FALLBACK_FORMULAS["downsell_1"],
        "downsell_2_price_fallback": DEFAULT_PRICE_FALLBACK_FORMULAS["downsell_2"],
        "upsell_1_price_fallback": DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_1"],
        "upsell_2_price_fallback": DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_2"],
    }


@pytest.fixture(autouse=True)
def disable_debug_toolbar(settings):
    settings.DEBUG_TOOLBAR_CONFIG = {
        **getattr(settings, "DEBUG_TOOLBAR_CONFIG", {}),
        "SHOW_TOOLBAR_CALLBACK": lambda request: False,
    }


@pytest.mark.django_db
def test_formula_lab_view_previews_custom_results(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    create_level("Highlighted", 1, 2, 2, "100.00")
    create_level("Default Downsell", 1, 2, 0, "80.00")
    create_level("Custom Downsell", 2, 2, 0, "90.00")
    create_level("Fallback Downsell", 1, 1, 2, "70.00")
    create_level("Upsell", 1, 3, 2, "120.00")

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "preview",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "(C+1, G, prev(T))",
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            **default_fallback_post_data(),
        },
    )

    assert response.status_code == 200
    assert response.context["match_type"] == "Exact"
    assert response.context["recommendation_rows"][1]["level"].name == "Custom Downsell"
    assert (
        response.context["recommendation_rows"][1]["formula"]
        == "(C+1, G, prev(T))"
    )
    assert b"Custom Downsell" in response.content
    assert b"(C+1, G, prev(T))" in response.content
    assert b"Show formula lab documentation" in response.content


@pytest.mark.django_db
def test_formula_lab_view_previews_custom_price_fallback_results(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    create_level("Highlighted", 1, 2, 2, "100.00")
    create_level("Cheaper First", 1, 0, 0, "70.00")
    create_level("Cheaper Second", 1, 2, 0, "90.00")
    create_level("Upsell", 1, 3, 2, "120.00")

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "preview",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "",
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            "downsell_1_price_fallback": "cheaper(1; match=cardholders,guests)",
            **{
                key: value
                for key, value in default_fallback_post_data().items()
                if key != "downsell_1_price_fallback"
            },
        },
    )

    assert response.status_code == 200
    assert response.context["recommendation_rows"][1]["level"].name == "Cheaper Second"
    assert (
        response.context["recommendation_rows"][1]["formula"]
        == "cheaper(1; match=cardholders,guests)"
    )
    assert b"Cheaper Second" in response.content
    assert b"cheaper(1; match=cardholders,guests)" in response.content


@pytest.mark.django_db
def test_formula_lab_view_shows_validation_errors(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "preview",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "(C, G)",
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            **default_fallback_post_data(),
        },
    )

    assert response.status_code == 200
    assert response.context["recommendation_rows"] == []
    assert (
        "Formula must contain exactly three expressions."
        in response.context["form"].errors["downsell_1_formulas"][0]
    )


@pytest.mark.django_db
def test_formula_lab_view_shows_price_fallback_validation_errors(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "preview",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_1"]
            ),
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            "downsell_1_price_fallback": "cheaper(1; match=color)",
            **{
                key: value
                for key, value in default_fallback_post_data().items()
                if key != "downsell_1_price_fallback"
            },
        },
    )

    assert response.status_code == 200
    assert response.context["recommendation_rows"] == []
    assert (
        "Fallback dimensions must be cardholders, guests, tickets, or their aliases C, G, T."
        in response.context["form"].errors["downsell_1_price_fallback"][0]
    )


@pytest.mark.django_db
def test_formula_lab_view_downloads_fixture_backed_matrix(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "download_matrix",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "(C+1, G, prev(T))",
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            "downsell_1_price_fallback": "cheaper(1; match=cardholders,guests)",
            **{
                key: value
                for key, value in default_fallback_post_data().items()
                if key != "downsell_1_price_fallback"
            },
        },
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response["Content-Disposition"]
        == 'attachment; filename="membership_matrix_formulas.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Matrix", "Levels", "Meta"]

    meta = {
        key: value
        for key, value in workbook["Meta"].iter_rows(values_only=True)
        if key is not None
    }
    assert meta["level_source"] == str(DEFAULT_MEMBERSHIP_MATRIX_FIXTURE_PATH)
    assert meta["downsell_1_formulas"] == "(C+1, G, prev(T))"
    assert meta["downsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
    )
    assert meta["upsell_1_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
    )
    assert meta["upsell_2_formulas"] == "\n".join(
        DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
    )
    assert meta["downsell_1_price_fallback"] == "cheaper(1; match=cardholders,guests)"
    assert (
        meta["downsell_2_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["downsell_2"]
    )
    assert (
        meta["upsell_1_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_1"]
    )
    assert (
        meta["upsell_2_price_fallback"]
        == DEFAULT_PRICE_FALLBACK_FORMULAS["upsell_2"]
    )


@pytest.mark.django_db
def test_formula_lab_view_invalid_download_rerenders_with_errors(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "download_matrix",
            "cardholders": 1,
            "admissions": 2,
            "member_tickets": 2,
            "downsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_1"]
            ),
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            "downsell_1_price_fallback": "cheaper(1; match=color)",
            **{
                key: value
                for key, value in default_fallback_post_data().items()
                if key != "downsell_1_price_fallback"
            },
        },
    )

    assert response.status_code == 200
    assert "Content-Disposition" not in response.headers
    assert response.context["recommendation_rows"] == []
    assert (
        "Fallback dimensions must be cardholders, guests, tickets, or their aliases C, G, T."
        in response.context["form"].errors["downsell_1_price_fallback"][0]
    )


@pytest.mark.django_db
def test_formula_lab_view_download_ignores_invalid_preview_request(client, settings):
    settings.MEMBERS_BASIC_AUTH_ENABLED = False

    response = client.post(
        reverse("members:membership_formula_lab"),
        data={
            "action": "download_matrix",
            "cardholders": 1,
            "admissions": 0,
            "member_tickets": 6,
            "downsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_1"]
            ),
            "downsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["downsell_2"]
            ),
            "upsell_1_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_1"]
            ),
            "upsell_2_formulas": "\n".join(
                DEFAULT_RECOMMENDATION_FORMULAS["upsell_2"]
            ),
            **default_fallback_post_data(),
        },
    )

    assert response.status_code == 200
    assert "Content-Disposition" not in response.headers
    assert response.context["recommendation_rows"] == []
    assert response.context["form"].non_field_errors()
