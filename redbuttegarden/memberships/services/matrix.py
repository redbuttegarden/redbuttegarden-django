from __future__ import annotations

import json
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping, Sequence

from memberships.forms import MembershipSelectorForm
from memberships.services.recommendations import (
    RECOMMENDATION_SLOT_ORDER,
    Level,
    get_default_price_fallback_formulas,
    get_default_recommendation_formulas,
    recommend_levels,
    validate_price_fallback_formula,
    validate_recommendation_formula,
)

DEFAULT_MEMBERSHIP_MATRIX_FIXTURE_PATH = (
    Path(__file__).resolve().parents[1] / "fixtures" / "membership_levels.json"
)
MATRIX_HEADERS = [
    "cardholders",
    "admissions",
    "tickets",
    "valid",
    "error",
    "match_type",
    "highlighted_formula",
    "highlighted_name",
    "highlighted_price",
    "highlighted_ticket_allowance",
    "downsell_1_formula",
    "downsell_1_name",
    "downsell_1_price",
    "downsell_2_formula",
    "downsell_2_name",
    "downsell_2_price",
    "upsell_1_formula",
    "upsell_1_name",
    "upsell_1_price",
    "upsell_2_formula",
    "upsell_2_name",
    "upsell_2_price",
]
LEVEL_HEADERS = [
    "pk",
    "name",
    "cardholders_included",
    "admissions_allowed",
    "member_sale_ticket_allowance",
    "price",
    "active",
]
MATRIX_NOTES = (
    "Matrix built from MembershipSelectorForm domain + membership_levels.json "
    "fixture. Suggestion columns map to: downsell_1, downsell_2, upsell_1, "
    "upsell_2 (named slots)."
)


def load_levels_from_fixture(path: Path) -> list[Level]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    levels: list[Level] = []

    for obj in raw:
        model = (obj.get("model") or "").lower()
        if not model.endswith(".membershiplevel"):
            continue

        fields = obj["fields"]
        levels.append(
            Level(
                pk=int(obj["pk"]),
                name=fields["name"],
                cardholders_included=int(fields["cardholders_included"]),
                admissions_allowed=int(fields["admissions_allowed"]),
                member_sale_ticket_allowance=int(
                    fields["member_sale_ticket_allowance"]
                ),
                price=Decimal(str(fields["price"])),
                active=bool(fields.get("active", True)),
            )
        )

    return levels


def _normalize_formulas(
    formulas: Mapping[str, Sequence[str]] | None,
) -> dict[str, tuple[str, ...]]:
    normalized = get_default_recommendation_formulas()
    if not formulas:
        return normalized

    unknown_slots = set(formulas) - set(RECOMMENDATION_SLOT_ORDER)
    if unknown_slots:
        unknown = ", ".join(sorted(unknown_slots))
        raise ValueError(f"Unknown recommendation slots: {unknown}")

    for slot in RECOMMENDATION_SLOT_ORDER:
        if slot not in formulas:
            continue

        slot_formulas = tuple(
            formula.strip() for formula in formulas[slot] if formula and formula.strip()
        )
        for formula in slot_formulas:
            validate_recommendation_formula(formula)
        normalized[slot] = slot_formulas

    return normalized


def _normalize_price_fallbacks(
    price_fallbacks: Mapping[str, str] | None,
) -> dict[str, str]:
    normalized = get_default_price_fallback_formulas()
    if not price_fallbacks:
        return normalized

    unknown_slots = set(price_fallbacks) - set(RECOMMENDATION_SLOT_ORDER)
    if unknown_slots:
        unknown = ", ".join(sorted(unknown_slots))
        raise ValueError(f"Unknown recommendation slots: {unknown}")

    for slot in RECOMMENDATION_SLOT_ORDER:
        if slot not in price_fallbacks:
            continue

        formula = price_fallbacks[slot].strip()
        validate_price_fallback_formula(formula)
        normalized[slot] = formula

    return normalized


def build_membership_matrix_rows(
    levels: Sequence[Level],
    cfg,
    formulas: Mapping[str, Sequence[str]] | None = None,
    price_fallbacks: Mapping[str, str] | None = None,
) -> list[dict[str, Any]]:
    normalized_formulas = _normalize_formulas(formulas)
    normalized_price_fallbacks = _normalize_price_fallbacks(price_fallbacks)
    rows: list[dict[str, Any]] = []

    for cardholders in range(1, 4):
        for admissions in range(0, 9):
            for tickets in (0, 2, 4, 6):
                form = MembershipSelectorForm(
                    data={
                        "cardholders": cardholders,
                        "admissions": admissions,
                        "member_tickets": tickets,
                    },
                    cfg=cfg,
                )
                valid = form.is_valid()

                if not valid:
                    non_field_errors = list(form.non_field_errors())
                    field_errors: list[str] = []
                    for field_name in ("cardholders", "admissions", "member_tickets"):
                        errors = form.errors.get(field_name)
                        if errors:
                            field_errors.extend(
                                [f"{field_name}: {error}" for error in errors]
                            )
                    error_text = " | ".join(non_field_errors + field_errors)

                    rows.append(
                        {
                            "cardholders": cardholders,
                            "admissions": admissions,
                            "tickets": tickets,
                            "valid": False,
                            "error": error_text,
                            "match_type": None,
                            "highlighted_formula": None,
                            "highlighted_name": None,
                            "highlighted_price": None,
                            "highlighted_ticket_allowance": None,
                            "downsell_1_formula": None,
                            "downsell_1_name": None,
                            "downsell_1_price": None,
                            "downsell_2_formula": None,
                            "downsell_2_name": None,
                            "downsell_2_price": None,
                            "upsell_1_formula": None,
                            "upsell_1_name": None,
                            "upsell_1_price": None,
                            "upsell_2_formula": None,
                            "upsell_2_name": None,
                            "upsell_2_price": None,
                        }
                    )
                    continue

                recommendation = recommend_levels(
                    levels=levels,
                    cardholders=cardholders,
                    guests=admissions,
                    tickets=tickets,
                    formulas=normalized_formulas,
                    price_fallbacks=normalized_price_fallbacks,
                )
                highlighted = recommendation.highlighted
                downsell_1 = recommendation.downsell_1
                downsell_2 = recommendation.downsell_2
                upsell_1 = recommendation.upsell_1
                upsell_2 = recommendation.upsell_2

                rows.append(
                    {
                        "cardholders": cardholders,
                        "admissions": admissions,
                        "tickets": tickets,
                        "valid": True,
                        "error": "",
                        "match_type": recommendation.match_type,
                        "highlighted_formula": recommendation.highlighted_formula,
                        "highlighted_name": highlighted.name if highlighted else None,
                        "highlighted_price": (
                            str(highlighted.price) if highlighted else None
                        ),
                        "highlighted_ticket_allowance": (
                            highlighted.member_sale_ticket_allowance
                            if highlighted
                            else None
                        ),
                        "downsell_1_formula": recommendation.downsell_1_formula,
                        "downsell_1_name": (
                            downsell_1.name if downsell_1 else None
                        ),
                        "downsell_1_price": (
                            str(downsell_1.price) if downsell_1 else None
                        ),
                        "downsell_2_formula": recommendation.downsell_2_formula,
                        "downsell_2_name": (
                            downsell_2.name if downsell_2 else None
                        ),
                        "downsell_2_price": (
                            str(downsell_2.price) if downsell_2 else None
                        ),
                        "upsell_1_formula": recommendation.upsell_1_formula,
                        "upsell_1_name": upsell_1.name if upsell_1 else None,
                        "upsell_1_price": str(upsell_1.price) if upsell_1 else None,
                        "upsell_2_formula": recommendation.upsell_2_formula,
                        "upsell_2_name": upsell_2.name if upsell_2 else None,
                        "upsell_2_price": str(upsell_2.price) if upsell_2 else None,
                    }
                )

    return rows


def build_membership_matrix_workbook_bytes(
    rows: Sequence[Mapping[str, Any]],
    levels: Sequence[Level],
    level_source,
    formulas: Mapping[str, Sequence[str]] | None = None,
    price_fallbacks: Mapping[str, str] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    normalized_formulas = _normalize_formulas(formulas)
    normalized_price_fallbacks = _normalize_price_fallbacks(price_fallbacks)
    workbook = Workbook()

    matrix_sheet = workbook.active
    matrix_sheet.title = "Matrix"

    header_font = Font(bold=True)
    for column, header in enumerate(MATRIX_HEADERS, start=1):
        cell = matrix_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        matrix_sheet.append([row.get(header) for header in MATRIX_HEADERS])

    for column_index in range(1, len(MATRIX_HEADERS) + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in matrix_sheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        matrix_sheet.column_dimensions[column_letter].width = min(
            max(10, max_length + 2), 80
        )

    matrix_sheet.freeze_panes = "A2"

    levels_sheet = workbook.create_sheet("Levels")
    for column, header in enumerate(LEVEL_HEADERS, start=1):
        cell = levels_sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for level in sorted(levels, key=lambda item: item.pk):
        levels_sheet.append(
            [
                level.pk,
                level.name,
                level.cardholders_included,
                level.admissions_allowed,
                level.member_sale_ticket_allowance,
                str(level.price),
                level.active,
            ]
        )

    levels_sheet.freeze_panes = "A2"

    meta_sheet = workbook.create_sheet("Meta")
    meta_sheet.append(["level_source", str(level_source)])
    meta_sheet.append(["notes", MATRIX_NOTES])
    for slot in RECOMMENDATION_SLOT_ORDER:
        meta_sheet.append([f"{slot}_formulas", "\n".join(normalized_formulas[slot])])
    for slot in RECOMMENDATION_SLOT_ORDER:
        meta_sheet.append([f"{slot}_price_fallback", normalized_price_fallbacks[slot]])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
