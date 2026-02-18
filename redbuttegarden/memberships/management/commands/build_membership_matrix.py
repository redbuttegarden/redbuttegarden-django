from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from django.core.management.base import BaseCommand

from memberships.forms import MembershipSelectorForm
from memberships.widget_config import MembershipWidgetConfig
from memberships.services.recommendations import Level, recommend_levels


def load_levels_from_fixture(path: Path) -> List[Level]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    levels: List[Level] = []

    for obj in raw:
        model = (obj.get("model") or "").lower()
        if not model.endswith(".membershiplevel"):
            continue

        pk = int(obj["pk"])
        f = obj["fields"]
        levels.append(
            Level(
                pk=pk,
                name=f["name"],
                cardholders_included=int(f["cardholders_included"]),
                admissions_allowed=int(f["admissions_allowed"]),
                member_sale_ticket_allowance=int(f["member_sale_ticket_allowance"]),
                price=Decimal(str(f["price"])),
                active=bool(f.get("active", True)),
            )
        )

    return levels


class Command(BaseCommand):
    help = "Build an XLSX matrix of MembershipSelectorForm inputs -> recommendation outputs."

    def add_arguments(self, parser):
        parser.add_argument(
            "--fixture",
            default="/code/memberships/fixtures/membership_levels.json",
            help="Path to membership_levels.json fixture.",
        )
        parser.add_argument(
            "--out",
            default="membership_matrix.xlsx",
            help="Output xlsx filename (default: membership_matrix.xlsx).",
        )

    def handle(self, *args, **opts):
        fixture_path = Path(opts["fixture"])
        if not fixture_path.is_absolute():
            fixture_path = Path.cwd() / fixture_path
        if not fixture_path.exists():
            raise SystemExit(f"Fixture not found: {fixture_path}")

        out_path = Path(opts["out"])
        if not out_path.is_absolute():
            out_path = Path.cwd() / out_path

        levels = load_levels_from_fixture(fixture_path)

        # Use the same cfg as production so validation messages match.
        cfg = MembershipWidgetConfig.get_solo()

        # Input domain from MembershipSelectorForm
        cardholders_values = range(1, 4)  # 1..3
        admissions_values = range(0, 9)  # 0..8
        ticket_values = [0, 2, 4, 6]  # even only

        rows: List[Dict[str, Any]] = []

        for c in cardholders_values:
            for a in admissions_values:
                for t in ticket_values:
                    form = MembershipSelectorForm(
                        data={"cardholders": c, "admissions": a, "member_tickets": t},
                        cfg=cfg,
                    )
                    valid = form.is_valid()

                    if not valid:
                        nfe = list(form.non_field_errors())
                        field_errs: List[str] = []
                        for f_name in ("cardholders", "admissions", "member_tickets"):
                            errs = form.errors.get(f_name)
                            if errs:
                                field_errs.extend([f"{f_name}: {e}" for e in errs])
                        error_text = " | ".join(nfe + field_errs)

                        rows.append(
                            {
                                "cardholders": c,
                                "admissions": a,
                                "tickets": t,
                                "valid": False,
                                "error": error_text,
                                "match_type": None,
                                "highlighted_pk": None,
                                "highlighted_name": None,
                                "highlighted_price": None,
                                "highlighted_ticket_allowance": None,
                                "downsell_1_pk": None,
                                "downsell_1_name": None,
                                "downsell_1_price": None,
                                "downsell_2_pk": None,
                                "downsell_2_name": None,
                                "downsell_2_price": None,
                                "upsell_1_pk": None,
                                "upsell_1_name": None,
                                "upsell_1_price": None,
                                "upsell_2_pk": None,
                                "upsell_2_name": None,
                                "upsell_2_price": None,
                            }
                        )
                        continue

                    rec = recommend_levels(
                        levels=levels,
                        cardholders=c,
                        guests=a,
                        tickets=t,
                    )

                    highlighted = rec.highlighted
                    d1 = rec.downsell_1
                    d2 = rec.downsell_2
                    u1 = rec.upsell_1
                    u2 = rec.upsell_2

                    rows.append(
                        {
                            "cardholders": c,
                            "admissions": a,
                            "tickets": t,
                            "valid": True,
                            "error": "",
                            "match_type": rec.match_type,
                            "highlighted_pk": highlighted.pk if highlighted else None,
                            "highlighted_name": (
                                highlighted.name if highlighted else None
                            ),
                            "highlighted_price": (
                                str(highlighted.price) if highlighted else None
                            ),
                            "highlighted_ticket_allowance": (
                                highlighted.member_sale_ticket_allowance
                                if highlighted
                                else None
                            ),
                            "downsell_1_pk": d1.pk if d1 else None,
                            "downsell_1_name": d1.name if d1 else None,
                            "downsell_1_price": str(d1.price) if d1 else None,
                            "downsell_2_pk": d2.pk if d2 else None,
                            "downsell_2_name": d2.name if d2 else None,
                            "downsell_2_price": str(d2.price) if d2 else None,
                            "upsell_1_pk": u1.pk if u1 else None,
                            "upsell_1_name": u1.name if u1 else None,
                            "upsell_1_price": str(u1.price) if u1 else None,
                            "upsell_2_pk": u2.pk if u2 else None,
                            "upsell_2_name": u2.name if u2 else None,
                            "upsell_2_price": str(u2.price) if u2 else None,
                        }
                    )

        self._write_xlsx(out_path, rows, levels, fixture_path)

        self.stdout.write(self.style.SUCCESS(f"Wrote: {out_path}"))
        self.stdout.write(f"Rows: {len(rows)} | Levels loaded: {len(levels)}")

    def _write_xlsx(
        self,
        out_path: Path,
        rows: List[Dict[str, Any]],
        levels: List[Level],
        fixture_path: Path,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---- Sheet 1: Matrix ----
        ws = wb.active
        ws.title = "Matrix"

        headers = [
            "cardholders",
            "admissions",
            "tickets",
            "valid",
            "error",
            "match_type",
            "highlighted_pk",
            "highlighted_name",
            "highlighted_price",
            "highlighted_ticket_allowance",
            "downsell_1_pk",
            "downsell_1_name",
            "downsell_1_price",
            "downsell_2_pk",
            "downsell_2_name",
            "downsell_2_price",
            "upsell_1_pk",
            "upsell_1_name",
            "upsell_1_price",
            "upsell_2_pk",
            "upsell_2_name",
            "upsell_2_price",
        ]

        ws.append(headers)

        header_font = Font(bold=True)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([r.get(h) for h in headers])

        # Autosize columns (simple heuristic)
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

        ws.freeze_panes = "A2"

        # ---- Sheet 2: Levels ----
        ws2 = wb.create_sheet("Levels")
        ws2_headers = [
            "pk",
            "name",
            "cardholders_included",
            "admissions_allowed",
            "member_sale_ticket_allowance",
            "price",
            "active",
        ]
        ws2.append(ws2_headers)
        for col, h in enumerate(ws2_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for l in sorted(levels, key=lambda x: x.pk):
            ws2.append(
                [
                    l.pk,
                    l.name,
                    l.cardholders_included,
                    l.admissions_allowed,
                    l.member_sale_ticket_allowance,
                    str(l.price),
                    l.active,
                ]
            )

        ws2.freeze_panes = "A2"

        # ---- Sheet 3: Metadata ----
        ws3 = wb.create_sheet("Meta")
        ws3.append(["fixture_path", str(fixture_path)])
        ws3.append(
            [
                "notes",
                "Matrix built from MembershipSelectorForm domain + membership_levels.json fixture. "
                "Suggestion columns map to: downsell_1, downsell_2, upsell_1, upsell_2 (named slots).",
            ]
        )

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
